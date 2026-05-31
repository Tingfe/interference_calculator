import os
import tempfile
import unittest

from interference_calculator.gdms_import import (
    extract_profile_elements,
    label_to_isotope,
    parse_gdms_gdr_runs,
    parse_gdms_profile_gdr,
    parse_gdms_profile_file,
    parse_gdms_profile_trr,
    parse_gdms_profile_xlsx,
    parse_gdms_raw_runs,
    parse_gdms_trr_runs,
    summarize_profile,
    _profiles_from_trr_components,
    _trr_percent_abundance,
)


class GDMSImportTests(unittest.TestCase):
    def test_label_to_isotope_accepts_gdms_brace_labels(self):
        self.assertEqual(label_to_isotope('Fe{56}'), ('Fe', 56, '56Fe'))
        self.assertEqual(label_to_isotope(' U{238} '), ('U', 238, '238U'))
        self.assertIsNone(label_to_isotope('56Fe'))

    def test_summarize_profile_returns_apex_centroid_and_fwhm(self):
        summary = summarize_profile(
            [55.90, 55.92, 55.94, 55.96, 55.98],
            [0.0, 4.0, 10.0, 4.0, 0.0],
        )

        self.assertEqual(summary.point_count, 5)
        self.assertAlmostEqual(summary.apex_mz, 55.94)
        self.assertAlmostEqual(summary.apex_intensity, 10.0)
        self.assertAlmostEqual(summary.centroid_mz, 55.94)
        self.assertAlmostEqual(summary.fwhm, 1.0 / 30.0)

    def test_parse_gdms_profile_xlsx_extracts_profiles_and_elements(self):
        try:
            import openpyxl
        except ImportError as exc:
            raise unittest.SkipTest(f'openpyxl is optional: {exc}')

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'Fe{56}'
        sheet['A2'] = 'Mass'
        sheet['B2'] = 'Values'
        sheet['C2'] = 'Peaks'
        sheet['D1'] = 'U{238}'
        sheet['D2'] = 'Mass'
        sheet['E2'] = 'Values'
        sheet['F2'] = 'Peaks'
        for row, (mass, value) in enumerate(
            [(55.90, 0), (55.92, 4), (55.94, 10), (55.96, 4), (55.98, 0)],
            start=3,
        ):
            sheet.cell(row, 1).value = mass
            sheet.cell(row, 2).value = value
            sheet.cell(row, 4).value = mass + 182.0
            sheet.cell(row, 5).value = value / 2.0

        handle = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        handle.close()
        try:
            workbook.save(handle.name)
            profiles = parse_gdms_profile_xlsx(handle.name)
        finally:
            workbook.close()
            os.remove(handle.name)

        self.assertEqual([profile.label for profile in profiles], ['Fe{56}', 'U{238}'])
        self.assertEqual([profile.isotope for profile in profiles], ['56Fe', '238U'])
        self.assertEqual(extract_profile_elements(profiles), ['Fe', 'U'])
        self.assertAlmostEqual(profiles[0].centroid_mz, 55.94)
        self.assertEqual(
            profiles[0].profile_points,
            ((55.90, 0.0), (55.92, 4.0), (55.94, 10.0), (55.96, 4.0), (55.98, 0.0)),
        )

    def test_parse_gdms_profile_file_dispatches_excel_exports(self):
        try:
            import openpyxl
        except ImportError as exc:
            raise unittest.SkipTest(f'openpyxl is optional: {exc}')

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'As{75}'
        sheet['A2'] = 'Mass'
        sheet['B2'] = 'Values'
        sheet.cell(3, 1).value = 74.92
        sheet.cell(3, 2).value = 1.0

        handle = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        handle.close()
        try:
            workbook.save(handle.name)
            profiles = parse_gdms_profile_file(handle.name)
        finally:
            workbook.close()
            os.remove(handle.name)

        self.assertEqual([profile.label for profile in profiles], ['As{75}'])

    def test_trr_component_mapping_prefers_gdms_centroid_mass(self):
        component = {
            'values': {
                'm_ComponentName': 'Fe{56}',
                'm_Mass': {
                    'kind': 'array',
                    'values': [55.90, 55.92, 55.94, 55.96, 55.98],
                },
                'm_Current': {
                    'kind': 'array',
                    'values': [0.0, 4.0, 10.0, 4.0, 0.0],
                },
                'm_ICValues': {
                    'kind': 'array',
                    'values': [0.0, 0.0, 0.0, 0.0, 0.0],
                },
                'm_CentroidMassValue': 55.945,
                'm_PeakValue': 10.0,
                'm_Abundancy': 91.754,
            },
        }

        profiles = _profiles_from_trr_components([component], {})

        self.assertEqual([profile.label for profile in profiles], ['Fe{56}'])
        self.assertAlmostEqual(profiles[0].centroid_mz, 55.945)
        self.assertAlmostEqual(profiles[0].natural_abundance, 0.91754)
        self.assertAlmostEqual(profiles[0].apex_mz, 55.94)
        self.assertAlmostEqual(profiles[0].fwhm, 1.0 / 30.0)
        self.assertEqual(
            profiles[0].profile_points,
            ((55.90, 0.0), (55.92, 4.0), (55.94, 10.0), (55.96, 4.0), (55.98, 0.0)),
        )

    def test_trr_abundance_values_are_percentages(self):
        self.assertAlmostEqual(_trr_percent_abundance(91.754), 0.91754)
        self.assertAlmostEqual(_trr_percent_abundance(0.282), 0.00282)
        self.assertIsNone(_trr_percent_abundance(None))

    def test_parse_gdms_profile_trr_sample_when_available(self):
        sample = os.environ.get('GDMS_TRR_SAMPLE')
        if not sample or not os.path.exists(sample):
            raise unittest.SkipTest('set GDMS_TRR_SAMPLE to validate a real TRR file')

        profiles = parse_gdms_profile_trr(sample)
        runs = parse_gdms_trr_runs(sample)

        self.assertGreaterEqual(len(profiles), 1)
        self.assertGreaterEqual(len(runs), 1)
        self.assertEqual(runs[0].name, 'Run1')
        self.assertEqual(runs[0].sample_id, '1264A')
        labels = [profile.label for profile in profiles]
        self.assertIn('Fe{56}', labels)
        fe56 = profiles[labels.index('Fe{56}')]
        self.assertEqual(fe56.isotope, '56Fe')
        self.assertGreater(fe56.point_count, 0)
        self.assertAlmostEqual(fe56.centroid_mz, 55.94444720012818)
        self.assertAlmostEqual(fe56.natural_abundance, 0.91754)
        self.assertGreater(fe56.fwhm, 0)

    def test_parse_gdms_profile_gdr_sample_when_available(self):
        sample = os.environ.get('GDMS_GDR_SAMPLE')
        if not sample or not os.path.exists(sample):
            raise unittest.SkipTest('set GDMS_GDR_SAMPLE to validate a real GDR file')

        runs = parse_gdms_gdr_runs(sample)
        profiles = parse_gdms_profile_gdr(sample)
        raw_runs = parse_gdms_raw_runs(sample)

        self.assertEqual(len(runs), 5)
        self.assertEqual(len(raw_runs), 5)
        self.assertEqual(runs[0].name, 'Run1')
        self.assertEqual(runs[0].sample_id, '1264')
        self.assertEqual(len(runs[0].profiles), 53)
        self.assertEqual(len(profiles), 53)

        signatures = {
            tuple(sorted(profile.label for profile in run.profiles))
            for run in runs
        }
        self.assertEqual(len(signatures), 1)
        labels = [profile.label for profile in profiles]
        self.assertIn('Fe{56}', labels)
        self.assertIn('Bi{209}', labels)
        fe56 = profiles[labels.index('Fe{56}')]
        self.assertEqual(fe56.point_count, 1000)
        self.assertAlmostEqual(fe56.natural_abundance, 0.91754)
        self.assertGreater(fe56.fwhm, 0)


if __name__ == '__main__':
    unittest.main()
